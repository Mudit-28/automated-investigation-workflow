import os
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


HEADER_BG      = "1F4E79"   # Dark Blue
HEADER_FG      = "FFFFFF"   # White
UPI_HIT_BG     = "FFD966"   # Yellow  — UPI ID found
GATEWAY_HIT_BG = "C6EFCE"   # Green   — Gateway detected, no UPI
QR_HIT_BG      = "F4B942"   # Orange  — QR / Payment initiation endpoint
COLLECT_HIT_BG = "FF6B6B"   # Red     — UPI collect request detected
AGGREGATOR_BG  = "D9B3FF"   # Purple  — Aggregator hosted page, Merchant UPI hidden


class Storage:
    def __init__(self, output_dir="output/reports"):
        self.output_dir = output_dir
        os.makedirs(self.output_dir, exist_ok=True)
        self.workbook = openpyxl.Workbook()
        self.summary_sheet = self.workbook.active
        self.summary_sheet.title = "Summary"
        self.requests_sheet = self.workbook.create_sheet("Network Requests")
        self._setup_headers()

    def _setup_headers(self):
        summary_headers = [
            "Website URL", "Page Title", "Total Requests",
            "Payment API URLs", "UPI IDs Found", "Gateways Found",
            "Bank IFSC Codes", "Account Numbers",
            "Beneficiary Names",          
            "Screenshot Path", "Timestamp"
        ]
        summary_widths = [35, 25, 15, 60, 35, 35, 30, 30, 40, 45, 22]

        request_headers = [
            "Website URL", "Type", "Method", "API Endpoint",
            "Status", "Gateways Detected", "UPI IDs Found",
            "Beneficiary Names",          
            "Post Data", "Response Body", "QR Decoded", "Timestamp"
        ]
        request_widths = [30, 12, 10, 55, 10, 28, 28, 35, 40, 60, 45, 22]

        self._write_headers(self.summary_sheet, summary_headers, summary_widths)
        self._write_headers(self.requests_sheet, request_headers, request_widths)

        for sheet, headers in [
            (self.summary_sheet, summary_headers),
            (self.requests_sheet, request_headers)
        ]:
            sheet.freeze_panes = "A2"
            sheet.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"

    def _write_headers(self, sheet, headers, widths):
        header_font      = Font(bold=True, color=HEADER_FG)
        header_fill      = PatternFill("solid", fgColor=HEADER_BG)
        header_alignment = Alignment(horizontal="center", vertical="center")

        for col, (header, width) in enumerate(zip(headers, widths), 1):
            cell = sheet.cell(row=1, column=col, value=header)
            cell.font      = header_font
            cell.fill      = header_fill
            cell.alignment = header_alignment
            sheet.column_dimensions[get_column_letter(col)].width = width

        sheet.row_dimensions[1].height = 20

    def save_summary(self, summary):
        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        payment_urls = list(set([
            entry.get("url", "")
            for entry in summary.get("network_findings", [])
            if entry.get("url")
        ]))

        upi_ids  = summary.get("all_upi_ids", [])
        gateways = summary.get("all_gateways", [])

        row_data = [
            summary.get("url", ""),
            summary.get("page_title", ""),
            summary.get("total_payment_requests", 0),
            "\n".join(sorted(payment_urls)),
            "\n".join(summary.get("all_upi_ids", [])),
            "\n".join(summary.get("all_gateways", [])),
            "\n".join(summary.get("all_ifsc_codes", [])),
            "\n".join(summary.get("all_account_numbers", [])),
            "\n".join(summary.get("all_beneficiary_names", [])),
            summary.get("screenshot_path", ""),
            timestamp
        ]

        self.summary_sheet.append(row_data)
        row_num = self.summary_sheet.max_row

        self.summary_sheet.row_dimensions[row_num].height = 120
        for col in [4, 5, 6, 7, 8, 9]:
            self.summary_sheet.cell(row=row_num, column=col).alignment = Alignment(
                wrap_text=True, vertical="top"
            )

        if upi_ids:
            fill = PatternFill("solid", fgColor=UPI_HIT_BG)
            for col in range(1, 12):
                self.summary_sheet.cell(row=row_num, column=col).fill = fill

        print(f"[Storage] Summary row saved for: {summary.get('url', '')}")

    def save_requests(self, summary):
        timestamp   = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        website_url = summary.get("url", "")
        QR_KEYWORDS = ["qr-code", "qr/", "/qr", "initiate/qr", "payment/initiate"]

        for entry in summary.get("network_findings", []):
            upi_ids          = entry.get("upi_ids_found", [])
            gateways         = entry.get("gateways_found", [])
            beneficiary_names = entry.get("beneficiary_names", [])
            qr_decoded       = entry.get("qr_decoded", "") or ""
            is_collect       = entry.get("is_collect_request", False)
            aggregator       = entry.get("aggregator", None)
            is_qr = any(k in entry.get("url", "").lower() for k in QR_KEYWORDS)

            gateway_display = ", ".join(gateways)
            if aggregator:
                gateway_display += f" | Aggregator: {aggregator} — merchant UPI hidden"

            row_data = [
                website_url,
                entry.get("source", "").upper(),
                entry.get("method", ""),
                entry.get("url", ""),
                entry.get("status", ""),
                gateway_display,
                ", ".join(upi_ids),
                "\n".join(beneficiary_names),         
                entry.get("post_data", "") or "",      
                entry.get("response_body", "") or "",  
                entry.get("qr_decoded", "") or "",    
                timestamp                              
            ]

            self.requests_sheet.append(row_data)
            row_num = self.requests_sheet.max_row

            if is_collect:
                fill = PatternFill("solid", fgColor=COLLECT_HIT_BG)
            elif upi_ids:
                fill = PatternFill("solid", fgColor=UPI_HIT_BG)
            elif is_qr or qr_decoded:
                fill = PatternFill("solid", fgColor=QR_HIT_BG)
            elif aggregator:
                fill = PatternFill("solid", fgColor=AGGREGATOR_BG)
            elif gateways:
                fill = PatternFill("solid", fgColor=GATEWAY_HIT_BG)
            else:
                fill = None

            if fill:
                for col in range(1, 13):
                    self.requests_sheet.cell(row=row_num, column=col).fill = fill

            for col in [4, 8, 9, 10, 11]:
                self.requests_sheet.cell(row=row_num, column=col).alignment = Alignment(
                    wrap_text=True, vertical="top"
                )

            self.requests_sheet.row_dimensions[row_num].height = 40

        print(f"[Storage] {len(summary.get('network_findings', []))} request rows saved")

    def save(self, summary):
        try:
            self.save_summary(summary)
        except Exception as e:
            print(f"[Storage] Error saving summary: {e}")
        try:
            self.save_requests(summary)
        except Exception as e:
            print(f"[Storage] Error saving requests: {e}")
        self._write_file(summary)

    def _write_file(self, summary):
        from urllib.parse import urlparse
        parsed       = urlparse(summary.get("url", ""))
        website_name = parsed.netloc.replace(".", "_").replace("www_", "")
        timestamp    = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename     = f"{website_name}_investigation_{timestamp}.xlsx"
        filepath     = os.path.join(self.output_dir, filename)
        self.workbook.save(filepath)
        print(f"[Storage] Excel report saved: {filepath}")
        return filepath